"""
aggregate_results.py — Collect fold-level JSON results and compute mean ± std.

Reads all {dataset}_{task}_fold*.json from OUTPUT_DIR/results/,
groups by (dataset, task), computes statistics, and writes to an Excel file.

Usage:
    python aggregate_results.py
    python aggregate_results.py --output-dir ./output --data-dir data
"""

import os
import json
import glob
import argparse
import numpy as np
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ==============================================================================
# Metric definitions
# ==============================================================================
METRICS = ['acc', 'bacc', 'spec', 'sens', 'auc', 'f1']


def load_results(results_dir):
    """Load all JSON result files and group by (dataset, task)."""
    grouped = defaultdict(lambda: defaultdict(dict))  # {(dataset, task): {fold: {...}}}

    pattern = os.path.join(results_dir, "*.json")
    for json_path in sorted(glob.glob(pattern)):
        with open(json_path, 'r') as f:
            data = json.load(f)

        dataset = data['dataset']
        task = data['task']
        fold = data['fold']
        grouped[(dataset, task)][fold] = data

    return grouped


def compute_stats(metrics_list):
    """Compute mean ± std for a list of metric dicts."""
    if not metrics_list:
        return {}
    stats = {}
    for key in METRICS:
        values = [m.get(key, np.nan) for m in metrics_list]
        values = [v for v in values if not np.isnan(v)]
        if values:
            stats[key] = {
                'mean': np.mean(values),
                'std': np.std(values),
            }
        else:
            stats[key] = {'mean': np.nan, 'std': np.nan}
    return stats


def build_report(grouped):
    """Build report table from grouped results."""
    report = []
    for (dataset, task), folds in sorted(grouped.items()):
        metrics_list = list(folds.values())
        stats = compute_stats(metrics_list)
        row = {
            'dataset': dataset,
            'task': task,
            'num_folds': len(metrics_list),
            **{f"{k}_mean": stats[k]['mean'] for k in METRICS},
            **{f"{k}_std":  stats[k]['std']  for k in METRICS},
        }
        report.append(row)
    return report


def format_mean_std(mean_val, std_val, fmt="{:.4f} ± {:.4f}"):
    """Format a mean ± std string."""
    if np.isnan(mean_val):
        return "-"
    return fmt.format(mean_val, std_val)


def write_excel(report, output_path, template_path=None):
    """
    Write report to Excel. If template_path is provided, write into
    the existing layout of the 实验记录.xlsx template.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MAD-Former Results"

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Headers
    headers = ['Dataset', 'Task', 'Folds']
    for m in METRICS:
        headers.append(m.upper())

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    for r, row_data in enumerate(report, 2):
        values = [
            row_data['dataset'],
            row_data['task'],
            row_data['num_folds'],
        ]
        for m in METRICS:
            values.append(format_mean_std(row_data[f'{m}_mean'], row_data[f'{m}_std']))

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    # Column widths
    col_widths = [18, 18, 8] + [22] * len(METRICS)
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # ---- Summary sheet (mean only, compact) ----
    ws2 = wb.create_sheet("Summary")
    summary_headers = ['Dataset', 'Task'] + [m.upper() for m in METRICS]
    for col, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for r, row_data in enumerate(report, 2):
        ws2.cell(row=r, column=1, value=row_data['dataset']).border = thin_border
        ws2.cell(row=r, column=2, value=row_data['task']).border = thin_border
        for col, m in enumerate(METRICS, 3):
            cell = ws2.cell(row=r, column=col,
                            value=format_mean_std(row_data[f'{m}_mean'], row_data[f'{m}_std']))
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    for col, w in enumerate([18, 18] + [22] * len(METRICS), 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    wb.save(output_path)
    print(f"\nExcel report saved to: {output_path}")


def print_terminal(report):
    """Print a terminal-friendly table."""
    print("\n" + "=" * 120)
    print(" RESULTS SUMMARY")
    print("=" * 120)

    header = f"{'Dataset':<16} {'Task':<14} {'Folds':<6}"
    for m in METRICS:
        header += f" {m.upper():>22}"
    print(header)
    print("-" * 120)

    for row in report:
        line = f"{row['dataset']:<16} {row['task']:<14} {row['num_folds']:<6}"
        for m in METRICS:
            line += f" {format_mean_std(row[f'{m}_mean'], row[f'{m}_std']):>22}"
        print(line)
    print("=" * 120)


def main():
    parser = argparse.ArgumentParser(description='Aggregate MAD-Former experiment results')
    parser.add_argument('--output-dir', type=str, default='./output',
                        help='Root output directory containing results/')
    parser.add_argument('--data-dir', type=str, default='data',
                        help='Data directory (unused, for compatibility)')
    parser.add_argument('--excel', type=str, default=None,
                        help='Output Excel path (default: output/results/summary.xlsx)')
    args = parser.parse_args()

    results_dir = os.path.join(args.output_dir, "results")
    if not os.path.isdir(results_dir):
        print(f"[ERROR] Results directory not found: {results_dir}")
        print("Run experiments first: bash run_experiments.sh")
        return

    grouped = load_results(results_dir)
    if not grouped:
        print(f"[ERROR] No JSON result files found in {results_dir}")
        return

    report = build_report(grouped)
    print_terminal(report)

    excel_path = args.excel or os.path.join(results_dir, "summary.xlsx")
    write_excel(report, excel_path)


if __name__ == '__main__':
    main()
